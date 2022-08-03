import tweepy
import config
from openpyxl import Workbook, load_workbook
import pandas as pd
import os
import json
import dateutil.parser
from loguru import logger


class Listener(tweepy.StreamingClient):
    def on_connect(self):
        logger.info("Stream connected")

    def set_user_id(self, uid):
        self.uid = uid

    def on_data(self, raw_data):
        logger.debug(raw_data)
        tweets = json.loads(raw_data)
        users = {u["id"]: u["username"] for u in tweets["includes"]["users"]}
        sheet_name = "Replies"

        if self.uid == int(tweets["data"]["author_id"]):
            sheet_name = "Tweets"
            if (
                tweets["data"].get("referenced_tweets", [{}])[0].get("type")
                == "retweeted"
            ):
                sheet_name = "Retweets"

        wb_filename = "user_tweets_data.xlsx"
        wb = load_workbook(wb_filename)
        page = wb[sheet_name]
        page.append(
            [
                dateutil.parser.parse(tweets["data"]["created_at"]).strftime(
                    "%d/%m/%Y %H:%M:%S"
                ),
                users[tweets["data"]["author_id"]],
                f'{tweets["data"]["conversation_id"]}',
                tweets["data"]["text"],
                users[tweets["data"].get("in_reply_to_user_id")]
                if tweets["data"].get("in_reply_to_user_id")
                else "",
            ]
        )
        wb.save(filename=wb_filename)


client = tweepy.Client(bearer_token=config.BEARER_TOKEN, wait_on_rate_limit=True)


def parse_result(tweets):
    result = []

    users = {u["id"]: u["username"] for u in tweets.includes["users"]}

    for tweet in tweets.data:
        logger.debug(tweet)
        result.append(
            {
                "created_at": tweet.created_at.strftime("%d/%m/%Y %H:%M:%S"),
                "author": users[tweet.author_id],
                "conversation_id": f"{tweet.conversation_id}",
                "text": tweet.text,
                "in_reply_to": users.get(tweet.in_reply_to_user_id),
            }
        )
    logger.debug(result)
    return result


def get_all_tweets(uid, start_time):
    all_tweets = []
    next_token = None

    while True:
        tweets = client.get_users_tweets(
            id=uid,
            max_results=100,
            tweet_fields=["created_at", "conversation_id"],
            expansions=["author_id", "in_reply_to_user_id"],
            pagination_token=next_token,
            start_time=start_time,
        )
        all_tweets += parse_result(tweets)

        next_token = tweets.meta.get("next_token", False)

        if next_token == False:
            break

    logger.debug(all_tweets)
    return all_tweets


def get_all_replies(uid, start_time):
    all_tweets = []
    next_token = None

    while True:
        tweets = client.get_users_mentions(
            id=uid,
            max_results=100,
            tweet_fields=["created_at", "conversation_id"],
            expansions=["author_id", "in_reply_to_user_id"],
            start_time=start_time,
        )

        all_tweets += parse_result(tweets)

        next_token = tweets.meta.get("next_token", False)

        if next_token == False:
            break

    logger.debug(all_tweets)
    return all_tweets


def create_tweet_file():
    wb = Workbook()
    ws = wb.active
    ws.title = "Tweets"
    wb.save(export_filename)


def main(username, start_time, export_filename):
    uid = client.get_user(username=username).data.id
    stream = Listener(bearer_token=config.BEARER_TOKEN, wait_on_rate_limit=True)
    stream.set_user_id(uid)

    if not os.path.exists(export_filename):
        logger.debug("Creating file...")
        create_tweet_file()
        all_tweets_df = pd.DataFrame(get_all_tweets(uid, start_time))[::-1]
        user_tweets = all_tweets_df[~all_tweets_df["text"].str.startswith("RT")]
        user_retweets = all_tweets_df[all_tweets_df["text"].str.startswith("RT")]

        logger.debug("Saving data to file...")
        with pd.ExcelWriter(
            export_filename, mode="a", engine="openpyxl", if_sheet_exists="replace"
        ) as writer:
            user_tweets.to_excel(writer, sheet_name="Tweets", index=False)
            user_retweets.to_excel(writer, sheet_name="Retweets", index=False)
            pd.DataFrame(get_all_replies(uid, start_time))[::-1].to_excel(
                writer, sheet_name="Replies", index=False
            )

    stream.add_rules(tweepy.StreamRule(f"from:{username} OR to:{username}"))
    stream.filter(
        tweet_fields=["created_at", "conversation_id", "referenced_tweets"],
        expansions=["author_id", "in_reply_to_user_id"],
    )


if __name__ == "__main__":
    username = "scndryan"
    logger.info(f"Scraping for user: {username}")
    export_filename = "user_tweets_data.xlsx"

    start_time = "2021-09-01T00:00:00Z"
    main(username, start_time, export_filename)
